import json
import time
from collections import namedtuple
import openpyxl
import requests

from django.conf import settings
from django.core.management import BaseCommand
from django.template.loader import render_to_string
from requests.auth import HTTPBasicAuth

from web.models import Entry
from web.submissions_processing import merge_fields_with_submission_data


class Command(BaseCommand):
    help = "Exports to HumaneAI Wordpress"
    _taxonomy_cache = {}
    background_id = 187
    project_type_id = 15
    session = requests.Session()
    partners = {}
    wps = {
        "WP1": 24,
        "WP2": 25,
        "WP3": 26,
        "WP4": 27,
        "WP5": 28,
        "WP6": 29,
        "WP7": 30,
        "WP67": 54,
        "WP8": 31,
        "WP9": 32,
        "WP10": 33,
        "WP11": 34,
        "WP1WP2": 35,
    }

    def handle(self, *args, **options):
        self.session.auth = HTTPBasicAuth(
            settings.EXPORT_WP_USER, settings.EXPORT_WP_PASS
        )
        # self.session.verify = False

        self.partners = self._create_partners_dict()
        self._export_entries()

    def _get_or_create_taxonomy_item(self, slug, name, parent=None):
        key = f"{slug}_{name}_{parent}"
        if self._taxonomy_cache.get(key):
            return self._taxonomy_cache[key]

        response = self.session.get(
            settings.EXPORT_WP_TAXONOMY_URL, params={"slug": slug}
        ).json()

        if not response:
            params = {"slug": slug, "name": name}

            if parent:
                params["parent"] = parent

            response = self.session.post(
                settings.EXPORT_WP_TAXONOMY_URL,
                params=params,
            ).json()

            self._taxonomy_cache[key] = response["id"]

            return response["id"]
        else:
            self._taxonomy_cache[key] = response[0]["id"]
            return response[0]["id"]

    def _create_partners_dict(self):
        workbook = openpyxl.load_workbook("partners.xlsx")
        sheet = workbook.active

        header = []
        for col in range(1, sheet.max_column + 1):
            header.append(sheet.cell(row=1, column=col).value)
        data = []

        for row in range(2, sheet.max_row + 1):
            elm = {}
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=col).value
                if value:
                    value = value.strip()
                elm[header[col - 1]] = value
            data.append(elm)

        partners_data = {}
        for item in data:
            acronym = item.get("Acronym")
            if acronym:
                acronym = acronym.upper()
                partners_data[acronym] = {
                    "department": item["Department"],
                    "institution": item["Institution"],
                }

        return partners_data

    def _export_entries(self):
        excluded_labels = [
            "Project title",
        ]
        LabelOrder = namedtuple("LabelOrder", ["label", "type", "new_label"])
        for entry in Entry.objects.filter(project__id=3, is_active=True).order_by(
            "-id"
        )[:10]:
            response = self.session.get(
                "{}/wp/v2/search/".format(settings.EXPORT_HUMANEAI_WP_URL),
                params={"search": entry.title, "subtype[]": "project"},
            ).json()
            if response:
                post_id = response[0]["id"]
            else:
                post_id = None

            submission = merge_fields_with_submission_data(
                fields=entry.project.fields,
                data=entry.data,
                excluded_labels=excluded_labels,
            )

            order_of_labels = [
                LabelOrder("Tagline", "text", ""),
                LabelOrder(
                    "Project description (1000-4000 char): including how it addresses the call topics",
                    "textarea",
                    "",
                ),
                LabelOrder(
                    "List and description of tangible results", "textarea", "Output"
                ),
                LabelOrder("Overall responsible person", "name", "Primary Contact"),
                LabelOrder("Project Partners", "text", "Project Partners"),
            ]

            institutions = []
            institutions_items = {
                "type": "list",
                "label": "Project Partners",
                "inputs": [],
            }
            for item in submission:
                if item["type"] == "list":
                    d = {}
                    for _input in item["inputs"]:
                        d[_input["label"]] = _input["value"]

                    acronym = d["Institution Name"].upper()
                    try:
                        institution_name = self.partners[acronym]["institution"]
                    except KeyError:
                        # print("missing", acronym, entry.id, entry.title)
                        institution_name = d["Institution Name"]

                    if institution_name not in institutions:
                        institutions.append(institution_name)
                    institutions_items["inputs"].append(
                        {
                            "label": "",
                            "type": "list",
                            "value": f"{institution_name}, {d['Contact Person']}",
                        }
                    )

            # ic(institutions_items)

            workpackages = []
            for item in submission:
                if (
                    item["label"]
                    == "Topic(s): (which  topic(s) in the call are being addressed"
                ):
                    for _input in item["inputs"]:
                        value = _input["value"]
                        if "WP 1-2" in value:
                            workpackages += [self.wps["WP1WP2"]]
                        elif "WP3" in value:
                            workpackages += [self.wps["WP3"]]
                        elif "WP4" in value:
                            workpackages += [self.wps["WP4"]]
                        elif "WP5" in value:
                            workpackages += [self.wps["WP5"]]
                        elif "WP 6" in value:
                            workpackages += [self.wps["WP6"]]
                        elif "WP6&7" in value:
                            workpackages += [self.wps["WP67"]]
                        elif "WP8" in value:
                            workpackages += [self.wps["WP8"]]
                        else:
                            raise

            final_submission = []
            for label_order in order_of_labels:
                for item in submission:
                    if (
                        item["label"] == label_order.label
                        and item["type"] == label_order.type
                    ):
                        if label_order.new_label is not None:
                            item["label"] = label_order.new_label

                        if label_order.type == "name":
                            d = {}
                            for _input in item["inputs"]:
                                d[_input["label"]] = _input["value"]

                            item[
                                "value"
                            ] = f"{d['First']} {d['Last']}, {institutions[0]}"
                            del item["inputs"]
                            item["type"] = "text"

                        final_submission.append(item)

            final_submission.insert(-1, institutions_items)
            # ic(final_submission)

            content = render_to_string(
                "web/export/wp_humaneai.html", {"submission_data": final_submission}
            )
            # print(content)

            project_cat = [self.project_type_id, *workpackages]
            print(project_cat)

            post_data = {
                "title": entry.title,
                "project_tagline": ", ".join(institutions),
                "project_cat": ",".join([str(i) for i in project_cat]),
                "project_status": "running",
                "status": "publish",
                "section": json.dumps(
                    [
                        {
                            "field_53ee2df353fed": "",
                            "field_53f1f503200df": [
                                {
                                    "acf_fc_layout": "simple_content",
                                    "field_53f1f52e200e0": content,
                                }
                            ],
                        }
                    ]
                ),
            }

            if post_id:
                post_data["ID"] = post_id
                url = f"{settings.EXPORT_HUMANEAI_WP_URL}/wp/v2/project/{post_id}"
            else:
                url = f"{settings.EXPORT_HUMANEAI_WP_URL}/wp/v2/project"

            if not post_id:
                post_data["project_image"] = self.background_id

            response = self.session.post(
                url,
                data=post_data,
            )
            # print(response.content)
            time.sleep(1)
            # breakpoint()
